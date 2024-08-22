import logging
import math
import os
import shutil
import smtplib
import sys
import time
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from glob import glob
from sre_constants import SUCCESS

import openpyxl
import pandas as pd
import win32com.client as client
import xlwings as xw


class files_and_folders:
    def __init__(self, data, bot_ID):
        self.log_file_path = data["log_file_path"]
        self.project_files = data["project_folder_path"]
        self.to_email = data["developer_email"]
        self.support_email = data["support_email"]
        self.error_file_path = data["error_file_path"]
        self.email_from = data["bot_email"]
        self.bot_ID = bot_ID

        self.attachment_path = os.path.join(self.project_files, "variables.txt")

    def delete_files_from_folder(self, folder):
        """Use this method to delete all files from a specific folder"""
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
                self.log_to_file("Files Deleted Successfully")
            except Exception as e:
                self.send_email(
                    "There was an error on line: "
                    + str(e.__traceback__.tb_lineno)
                    + "\n"
                    "\n"
                    "Failed to delete %s. Reason: %s" % (file_path, e)
                )
                self.log_to_file("Error When Deleting Files")

    def rename_file(self, old_name, new_name):
        """Use this method to rename a file"""
        try:
            os.rename(old_name, new_name)
            self.log_to_file("File Renamed Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to rename file %s. Reason: %s" % (old_name, e)
            )
            self.log_to_file("Error When Renaming File")

    def delete_folder(self, dir_path):
        """Use this method to delete a folder and any files it contains"""
        try:
            shutil.rmtree(dir_path)
            self.log_to_file("Folder Deleted Successfully")
        except OSError as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Error: %s : %s" % (dir_path, e.strerror)
            )
            self.log_to_file("Error When Deleting Folder")

    def copy_file(self, original, target):
        """Use this method to create a copy of a file"""
        try:
            shutil.copyfile(original, target)
            self.log_to_file("File Copied Successfully")
        except OSError as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Error: %s : %s" % (original, e.strerror)
            )
            self.log_to_file("Error When Copying File")

    def get_excel_sheet_names(self, file_path):
        """Use this method to get the names of the sheets on an Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            print(wb.sheetnames)
            self.log_to_file("Sheet Name(s) Saved Successfully")
            return wb.sheetnames
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to get sheet names: %s. Reason: %s" % (file_path, e)
            )
            self.log_to_file("Error When Trying to Get Excel Sheet Names")

    def create_excel_from_filter(
        self, file_path, column, item_to_filter, new_file_path
    ):
        """Use this method to filter data from an Excel file and create a new Excel file with this Data"""
        """item_to_filter = word, number, phrase, etc, that you wish to filter by from existing Excel file, ex: Month, Name, Amount, etc"""
        try:
            df = pd.read_excel(file_path)
            df_tech = df.loc[df[column] == item_to_filter]
            print(df_tech)
            df_tech.to_excel(new_file_path, sheet_name="Filtered", index=False)
            self.log_to_file("Excel Created Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to create excel from filter: %s. Reason: %s" % (file_path, e)
            )
            self.log_to_file("Error When Creating Excel File from Filter")

    def add_image(self, image_path, excel_path, sheet_name, cell):
        """Use this method to add an image to Excel"""
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_path)
            sheet = wb.sheets[sheet_name]
            files_location = glob.glob(self.project_files + r"\*")
            sheet.pictures.add(image_path, link_to_file=True, anchor=sheet.range(cell))
            wb.save()
            wb.close()
            app.quit()
            self.log_to_file("Image Added Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to add image: %s. Reason: %s" % (image_path, e)
            )
            self.log_to_file("Error When Adding Image")

    def transfer_data_excel(self, source_file, dest_file, sheet_name, cell):
        """Transfer data to an existing Excel file"""
        try:
            df = pd.read_excel(source_file, index_col=0)
            app = xw.App(visible=False)
            wb = app.books.open(dest_file)
            ws = wb.sheets[sheet_name]
            ws.range(cell).options(index=True).value = df
            wb.save()
            wb.close()
            app.quit()
            self.log_to_file("Data Transferred Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to transfer data: %s. Reason: %s" % (source_file, e)
            )
            self.log_to_file("Error When Transferring Data to Excel")

    def save_downloaded_file_name(self, download_folder_path):
        """Save the name of a downloaded file as a variable"""
        try:
            folder = download_folder_path
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                self.reportname = file_path
            print(self.reportname)
            self.log_to_file("Downloaded File Name Saved Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to save file name: %s. Reason: %s" % (download_folder_path, e)
            )
            self.log_to_file("Error When Saving the Name of Downloaded File")

    def move_file(self, source_path, dest_path):
        """Move a file from one location to another"""
        moved = False
        try:
            src_path = source_path
            dst_path = dest_path
            shutil.move(src_path, dst_path)
            moved = True
            self.log_to_file("File Moved Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to move file: %s. Reason: %s" % (source_path, e)
            )
            self.log_to_file("Error When Moving File")
        return moved

    def download_unread_attachments(self, subfolder, download_path):
        """Use this method to read unread emails and download attachments from Outlook"""
        try:
            outlook = client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            inbox = outlook.GetDefaultFolder(6)
            cc013 = inbox.Folders.Item(subfolder)
            messages = cc013.Items
            if len(messages) == 0:
                print("There aren't any messages in this folder")
                exit()
            emails = []
            for message in messages:
                if message.Unread:
                    this_message = (
                        message.Subject,
                        message.SenderEmailAddress,
                        message.To,
                        message.Unread,
                        message.Senton.date(),
                        message.body,
                        message.Attachments,
                    )
                    emails.append(this_message)
            for email in emails:
                (
                    subject,
                    from_address,
                    to_address,
                    if_read,
                    date_sent,
                    body,
                    attachments,
                ) = email
                print(subject, to_address, body)
                if len(attachments) == 0:
                    print("No attachments")
                else:
                    for attachment in attachments:
                        print(attachment)
                        attachment.SaveAsFile(download_path + attachment.FileName)
                        print("Saved {0} attachments".format(len(attachments)))
            self.log_to_file("Attachments Downloaded Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to download attachment: %s. Reason: %s" % (subfolder, e)
            )
            self.log_to_file("Error When Downloading Attachments from Unread Emails")

    def download_all_attachments(self, subfolder, download_path):
        """Use this method to read all emails and download attachments from Outlook"""
        try:
            outlook = client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            inbox = outlook.GetDefaultFolder(6)
            cc013 = inbox.Folders.Item(subfolder)
            messages = cc013.Items
            if len(messages) == 0:
                print("There aren't any messages in this folder")
                exit()
            emails = []
            for message in messages:
                this_message = (
                    message.Subject,
                    message.SenderEmailAddress,
                    message.To,
                    message.Unread,
                    message.Senton.date(),
                    message.body,
                    message.Attachments,
                )
                emails.append(this_message)
            for email in emails:
                (
                    subject,
                    from_address,
                    to_address,
                    if_read,
                    date_sent,
                    body,
                    attachments,
                ) = email
                print(subject, to_address, body)
                if len(attachments) == 0:
                    print("No attachments")
                else:
                    for attachment in attachments:
                        print(attachment)
                        attachment.SaveAsFile(download_path + attachment.FileName)
                        print("Saved {0} attachments".format(len(attachments)))
            self.log_to_file("Attachments Downloaded Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to download attachment: %s. Reason: %s" % (subfolder, e)
            )
            self.log_to_file("Error When Downloading Attachments from Unread Emails")

    def get_file_names_in_folder(self, path):
        """Use this method to get a list with the names of all files in a folder"""
        try:
            dir_list = os.listdir(path)
            print(dir_list)
            self.log_to_file("File Names Saved Successfully")
            return dir_list

        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to get names of file in folder: %s. Reason: %s" % (path, e)
            )
            self.log_to_file("Error When Trying to Get All File Names in a Folder")

    def check_if_file_exists(self, path):
        """Use this method to check if file exists"""
        isFile = False
        try:
            isFile = os.path.isfile(path)
            if isFile:
                self.log_to_file(f"File Exists in path: {path}")
            else:
                self.log_to_file(f"File doesn't exists in path: {path}")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to check if file exists: %s. Reason: %s" % (path, e)
            )
            self.log_to_file("Error When Checking if a File Exists")
        return isFile

    def check_if_directory_exists(self, path):
        """Use this method to check if directory exists"""
        try:
            path = path
            isdir = os.path.isdir(path)
            print(isdir)
            self.log_to_file("Successfully Checked if Directory Exists")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to check if directory exists: %s. Reason: %s" % (path, e)
            )
            self.log_to_file("Error When Checking if Directory Exists")

    def wait_for_file_to_be_downloaded(self, file_path, wait_in_seconds=60):
        exists = False
        try:
            number_of_iterations = math.ceil(wait_in_seconds / 2)
            # Check if file has been downloaded
            count = 0
            while not exists and count < number_of_iterations:
                time.sleep(2)
                exists = self.check_if_file_exists(file_path)
                count += 1
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to wait for file to be downloaded: %s. Reason: %s" % (path, e)
            )
            self.log_to_file("Error When waiting for file to be downloaded")
        return exists

    def send_email(self, email_body, error_screenshot=None):
        try:
            self.get_environment_variables(self.attachment_path)
            print("Creating the email message...")
            message = MIMEMultipart()
            message["From"] = self.email_from

            if isinstance(self.to_email, str):
                message["To"] = self.to_email
            else:
                self.to_email = "; ".join(self.to_email)
                message["To"] = self.to_email

            rcpt = [self.to_email]
            cc = self.support_email
            message["Cc"] = cc
            rcpt.append(cc)

            message["Date"] = formatdate(localtime=True)
            subject = f"An error occurred on {self.bot_ID}"
            message["Subject"] = subject
            message.attach(MIMEText(email_body))

            # attach environment variables file
            with open(self.attachment_path, "rb") as file:
                part = MIMEApplication(
                    file.read(), Name=os.path.basename(self.attachment_path)
                )
            part[
                "Content-Disposition"
            ] = 'attachment; filename="%s"' % os.path.basename(self.attachment_path)
            message.attach(part)

            # Attach error screenshot if any
            if error_screenshot:
                with open(error_screenshot, "rb") as file:
                    part = MIMEApplication(
                        file.read(), Name=os.path.basename(error_screenshot)
                    )
                part[
                    "Content-Disposition"
                ] = 'attachment; filename="%s"' % os.path.basename(error_screenshot)
                message.attach(part)

            print("Email Message created")
            print("Start sending email message...")

            try:
                print("Start connecting to the SMTP service...")
                with smtplib.SMTP("relay.experian.com", 25, timeout=5.00) as smtp_obj:
                    print("Connected to the SMTP service...")
                    smtp_obj.set_debuglevel(2)
                    smtp_obj.sendmail(self.email_from, rcpt, message.as_string())
                    smtp_obj.quit()
                    print("Email sent successfully")
            except Exception as ex:
                print("Email message not sent")
                raise Exception("Not able to connect to SMTP service")

        except Exception as ex:
            error_line = sys.exc_info()[-1].tb_lineno
            print(
                "An error has occurred while sending the email, error details:"
                + str(ex)
                + ". Error line :"
                + str(error_line)
            )

    def log_to_file(self, log_message):
        """Use this method to save a .txt file with the logs of your script"""
        Log_Format = "%(asctime)s - %(message)s"
        logging.basicConfig(
            filename=self.log_file_path,
            filemode="a",
            format=Log_Format,
            level=logging.ERROR,
        )
        logger = logging.getLogger()
        logger.error(log_message)

    def get_environment_variables(self, variables_txt_path):
        """Use this method to save the environment variables being used, in a .txt file"""
        try:
            enviroment_variables = os.environ
            file = open(variables_txt_path, "w")
            file.write(str(enviroment_variables))
            file.close
            self.log_to_file("Environment Variables File Created Successfully")
        except Exception as e:
            self.send_email(
                "There was an error on line: " + str(e.__traceback__.tb_lineno) + "\n"
                "\n"
                "Failed to create environment variables file: %s. Reason: %s"
                % (variables_txt_path, e)
            )
            self.log_to_file("Error When Creating Environment Variables File")

    def clean_folders(self, file_list=[], pdf_directory=[], folders=[]):
        """This method is used to delete the files downloaded on the run"""
        try:
            self.log_to_file("Deleting files")
            if pdf_directory:
                for pdf_item in pdf_directory:
                    files_in_directory = os.listdir(pdf_item)
                    filtered_files = [
                        file for file in files_in_directory if file.endswith(".pdf")
                    ]
                    for file in filtered_files:
                        path_to_file = os.path.join(pdf_item, file)
                        os.remove(path_to_file)
                        self.log_to_file(f"Deleted file : {file}")
            if folders:
                for folder_item in folders:
                    if os.path.exists(folder_item):
                        files_in_folder = os.listdir(folder_item)
                        for file in files_in_folder:
                            path_to_file = os.path.join(folder_item, file)
                            os.remove(path_to_file)
                            self.log_to_file(f"Deleted file : {file}")
            if file_list:
                for file in file_list:
                    if os.path.isfile(file):
                        os.remove(file)
                        self.log_to_file(f"Deleted file : {file}")
        except Exception as error:
            self.send_email(
                "There was an error on line: "
                + str(error.__traceback__.tb_lineno)
                + 'Please review the below error message: "\n'
                "\n" + str(error) + "\n"
                "\n"
                "Thank you!"
            )
            self.log_to_file("Error When Creating Environment Variables File")

    def taskkill(self, process_name):
        """This is used to kill any process on the machine
        Args:
            process_name (str): The name of the process that you wish to kill
        """
        try:
            for proc in psutil.process_iter():
                try:
                    if process_name.lower() in proc.name().lower():
                        self.salesops.log_to_file(f"Killing {proc.name()}")
                        print(f"######## Killing {proc.name()} ########")
                        proc.kill()
                        time.sleep(3)
                except (
                    psutil.NoSuchProcess,
                    psutil.AccessDenied,
                    psutil.ZombieProcess,
                ):
                    pass
        except Exception:
            pass

    def get_html_email(self, file_path):
        """This method is used to extract the html from a txt file
        Args:
            file_path (str): Path to the txt where the html is saved
        Returns:
            file_data (str): The html of the extracted txt file
        """
        file_data = None
        try:
            self.log_to_file(f"Getting HTML Body from : {file_path}")
            if os.path.isfile(file_path):
                with open(file_path, "r") as opened_file:
                    self.log_to_file("Extracted HTML for email")
                    file_data = opened_file.read()
            else:
                raise Exception(f"File not found. File: {file_path}")
        except Exception as error:
            self.send_email(
                "There was an error on line: "
                + str(error.__traceback__.tb_lineno)
                + 'Please review the below error message: "\n'
                "\n" + str(error) + "\n"
                "\n"
                "Thank you!"
            )
            self.log_to_file("Error When Creating Environment Variables File")
        return file_data
