# ----------------------------------------------------------------------------------------------

# Módulo para trabajar con programación asíncrona.
import asyncio

# Módulo para trabajar con datos en formato JSON.
import json

# Módulo para interactuar con el sistema operativo.
import os

# Módulo que proporciona funciones relacionadas con el manejo del tiempo.
import time

# Biblioteca utilizada para controlar un navegador web.
import pyppeteer

# Función que carga variables de entorno desde un archivo en el entorno de ejecución.
from dotenv import load_dotenv
from openpyxl import load_workbook

# Importa una utilidad personalizada para gestionar archivos y carpetas desde una biblioteca local o interna.
from libs.files_and_folders_utils import files_and_folders

# Importa una clase o extensión personalizada para Pyppeteer desde una biblioteca interna.
from libs.pyppeteer_class import pyppeteer_extension

# Función de la biblioteca OpenPyXL para trabajar con archivos Excel.

# ----------------------------------------------------------------------------------------------


# ----------------------------------------------------------------------------------------------


# Clase que representa un bot para interactuar con la aplicación Koyfin, permitiendo automatizar tareas relacionadas con el análisis de acciones y datos financieros.
class StocksBot:

    # ----------------------------------------------------------------------------------------------

    def __init__(self, json_file_path):
        try:
            # Intenta abrir y cargar los datos del archivo JSON de configuración.
            with open(json_file_path) as read_file:
                # Carga el archivo JSON en un diccionario llamado `data`.
                data = json.load(read_file)
        except FileNotFoundError:
            # Si el archivo de configuración no se encuentra, imprime un mensaje de error.
            print(f"Configuration file not found: {json_file_path}")
            return
        except json.JSONDecodeError:
            # Si hay un error al decodificar el JSON, imprime un mensaje de error.
            print(f"Error decoding JSON from the file: {json_file_path}")
            return

        # Carga las credenciales de inicio de sesión desde las variables de entorno definidas en el archivo .env.
        self.login_email = os.getenv("login_email")
        self.login_password = os.getenv("login_password")

        # Asigna los valores de configuración obtenidos del archivo JSON a variables de instancia.
        self.file_name = data.get("file_name")
        chrome_executable_path = data.get("chrome_executable_path")

        # Obtiene el bucle de eventos asíncrono actual y establece un manejador de excepciones personalizado.
        loop = asyncio.get_event_loop()
        loop.set_exception_handler(self.custom_exception_handler)

        # Inicializa la utilidad de manejo de archivos y carpetas con los datos del archivo JSON y el nombre del bot.
        self.utils = files_and_folders(data, "StocksBot")

        try:
            # Intenta ejecutar un navegador Chrome utilizando Pyppeteer.
            self.browser = loop.run_until_complete(
                pyppeteer.launch(
                    headless=False,
                    executablePath=chrome_executable_path,
                    args=["--start-maximized"],
                )
            )
        except pyppeteer.errors.BrowserError as e:
            # Captura y muestra un mensaje de error si no se puede ejecutar el navegador.
            print(f"Error launching browser: {e}")
            return

        # Inicializa la extensión de Pyppeteer, pasando el navegador y la ruta del archivo de registro.
        self.pyppeteer_extensions = pyppeteer_extension(
            self.browser, self.utils.log_file_path
        )

        # Llama al método `start_automation` para comenzar el proceso de automatización.
        self.start_automation()

    # ----------------------------------------------------------------------------------------------

    def custom_exception_handler(self, loop, context):
        # Muestra un mensaje personalizado cuando ocurre una excepción dentro del bucle de eventos asíncrono.
        print(f"Exception in event loop: {context['message']}")

    # ----------------------------------------------------------------------------------------------

    def login(self):
        try:
            # Abre la página de inicio de sesión de Koyfin y espera a que el campo de email esté presente.
            self.pyppeteer_extensions.open_web_page(
                "https://app.koyfin.com/login", '//input[@type="email"]'
            )

            # Introduce el email de inicio de sesión en el campo correspondiente.
            self.pyppeteer_extensions.set_text(
                '//input[@type="email"]', self.login_email
            )

            # Espera 0.5 segundos.
            time.sleep(0.5)

            # Introduce la contraseña de inicio de sesión en el campo correspondiente.
            self.pyppeteer_extensions.set_text(
                '//input[@name="password"]', self.login_password
            )

            # Hace clic en el botón de "Sign in" para enviar el formulario de inicio de sesión.
            self.pyppeteer_extensions.click('//label[text()="Sign in"]')

            # Espera a que aparezca el elemento que indica que el inicio de sesión fue exitoso.
            self.pyppeteer_extensions.wait_for_element(
                '//div[text()="Search for a name, ticker, or function "]'
            )

            # Verifica si existe un botón de "Accept All" (para aceptar cookies o políticas).
            if self.pyppeteer_extensions.element_exist('//button[text()="Accept All"]'):
                # Si el botón existe, hace clic en él para aceptarlo.
                self.pyppeteer_extensions.click('//button[text()="Accept All"]')

            # Devuelve True si el inicio de sesión fue exitoso.
            return True
        except Exception as e:
            # Captura cualquier excepción y muestra un mensaje de error en caso de que falle el proceso de inicio de sesión.
            print(f"Login failed: {e}")
            return False

    # ----------------------------------------------------------------------------------------------

    def start_automation(self):
        # Inicializa la variable `char` con el valor ASCII correspondiente a la letra 'C'.
        char = 67

        # Llama al método `get_ticker_list` para obtener una lista de tickers (símbolos de acciones) a procesar.
        tickers_list = self.get_ticker_list()

        # Verifica si la lista de tickers está vacía.
        if not tickers_list:
            # Imprime un mensaje si no se encuentran tickers y termina la ejecución del método.
            print("No tickers found.")
            return

        # Intenta iniciar sesión y almacena el resultado.
        login_success = self.login()

        # Verifica si el inicio de sesión fue exitoso.
        if not login_success:
            # Si el inicio de sesión falló, termina la ejecución del método.
            return

        # Itera a través de cada ticker en la lista de tickers.
        for ticker in tickers_list:
            try:
                # Llama al método `collect_and_write_data` para recopilar datos del ticker y escribirlos (probablemente en un archivo).
                self.collect_and_write_data(ticker, char)

                # Incrementa el valor de `char` para el siguiente ticker.
                char += 1
            except Exception as e:
                # Captura y muestra un mensaje de error si ocurre un problema al procesar el ticker.
                print(f"Error processing ticker {ticker}: {e}")

    # ----------------------------------------------------------------------------------------------

    def collect_and_write_data(self, ticker, char):
        pe_ratio = ""
        pb_ratio = ""
        stock_price = ""
        current_ratio = ""
        return_on_equity = ""
        asset_turnover = ""
        eps_current_fiscal_year = ""
        eps_last_fiscal_year = ""
        total_liabilities_total_assets = ""

        # Hace clic en el elemento que permite buscar por nombre, ticker o función en la interfaz de Koyfin.
        self.pyppeteer_extensions.click(
            '//div[text()="Search for a name, ticker, or function "]'
        )

        # Espera a que aparezca el campo de entrada donde se puede escribir el nombre o ticker.
        self.pyppeteer_extensions.wait_for_element(
            '//input[@placeholder="Type by name or ticker"]'
        )

        # Establece el texto del ticker en el campo de búsqueda.
        self.pyppeteer_extensions.set_text(
            '//input[@placeholder="Type by name or ticker"]', ticker
        )

        # Espera a que aparezca el primer elemento correspondiente al ticker buscado.
        self.pyppeteer_extensions.wait_for_element(
            f'(//div[@class="rc-dialog-content"]/descendant::div[text()="{ticker}"])[1]'
        )

        # Hace clic en el primer resultado que corresponde al ticker buscado.
        self.pyppeteer_extensions.click(
            f'(//div[@class="rc-dialog-content"]/descendant::div[text()="{ticker}"])[1]'
        )

        # Espera a que se cargue la sección de "Overview" (Resumen) del ticker.
        self.pyppeteer_extensions.wait_for_element(
            '//div[@class="console-popup__functionTitle___wZJdU"][text()="Overview"]'
        )

        # Hace clic en la pestaña "Overview" para mostrar información general del ticker.
        self.pyppeteer_extensions.click(
            '//div[@class="console-popup__functionTitle___wZJdU"][text()="Overview"]'
        )

        # Espera a que aparezca el elemento que contiene la relación Precio/Ganancias (P/E).
        self.pyppeteer_extensions.wait_for_element(
            '//div[text()="P/E"]/following-sibling::div/div/div'
        )

        # Obtiene el valor de P/E y lo almacena en pe_ratio.
        while not pe_ratio:
            pe_ratio = self.pyppeteer_extensions.get_text(
                '//div[text()="P/E"]/following-sibling::div/div/div'
            )

        # Espera a que aparezca el elemento que contiene la relación Precio/Valor Libro (P/B).
        self.pyppeteer_extensions.wait_for_element(
            '//div[text()="Price/Book"]/following-sibling::div/div/div'
        )

        # Obtiene el valor de P/B y lo almacena en pb_ratio.
        while not pb_ratio:
            pb_ratio = self.pyppeteer_extensions.get_text(
                '//div[text()="Price/Book"]/following-sibling::div/div/div'
            )

        # Espera a que aparezca el elemento que muestra el precio de la acción.
        self.pyppeteer_extensions.wait_for_element(
            f'//span[text()="{ticker}"]/parent::div/div/label'
        )

        # Obtiene el precio de la acción y lo almacena en stock_price.
        while not stock_price:
            stock_price = self.pyppeteer_extensions.get_text(
                f'//span[text()="{ticker}"]/parent::div/div/label'
            )

        # Espera 2 segundos para que la interfaz cargue completamente.
        time.sleep(2)

        # Espera a que aparezca el botón de análisis financiero.
        self.pyppeteer_extensions.wait_for_element(
            '//div[text()="Financial Analysis"]/parent::div/parent::div'
        )

        # Hace clic en el botón para abrir el análisis financiero.
        self.pyppeteer_extensions.click(
            '//div[text()="Financial Analysis"]/parent::div/parent::div'
        )

        # Espera 1 segundo para permitir que la nueva sección cargue.
        time.sleep(1)

        # Hace clic en la pestaña de "Profitability" (Rentabilidad) en el análisis financiero.
        self.pyppeteer_extensions.click('//div[text()="Profitability"]')

        # Espera a que aparezca el elemento que contiene la relación corriente.
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Current Ratio"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
        )

        # Obtiene el valor de la relación corriente y lo almacena en current_ratio.
        while not current_ratio:
            current_ratio = self.pyppeteer_extensions.get_text(
                '(//div[text()="Current Ratio"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
            )

        # Espera a que aparezca el elemento que contiene el retorno sobre el capital (ROE).
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Return On Equity"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
        )

        # Obtiene el valor de retorno sobre el capital y lo almacena en return_on_equity.
        while not return_on_equity:
            return_on_equity = self.pyppeteer_extensions.get_text(
                '(//div[text()="Return On Equity"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
            )

        # Espera a que aparezca el elemento que contiene la rotación de activos.
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Asset Turnover"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
        )

        # Obtiene el valor de la rotación de activos y lo almacena en asset_turnover.
        while not asset_turnover:
            asset_turnover = self.pyppeteer_extensions.get_text(
                '(//div[text()="Asset Turnover"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
            )

        # Espera 1 segundo para permitir que la interfaz cargue.
        time.sleep(1)

        # Hace clic en la pestaña "Highlights" (Aspectos Destacados) para obtener más información.
        self.pyppeteer_extensions.click('//span[text()="Highlights"]')

        # Espera a que aparezca el elemento que contiene el EPS diluido del año fiscal actual.
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Diluted EPS"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
        )

        # Obtiene el EPS del año fiscal actual y lo almacena en eps_current_fiscal_year.
        while not eps_current_fiscal_year:
            eps_current_fiscal_year = self.pyppeteer_extensions.get_text(
                '(//div[text()="Diluted EPS"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
            )

        # Espera a que aparezca el elemento que contiene el EPS del año fiscal anterior.
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Diluted EPS"]/parent::div/parent::div/parent::div/parent::div/div)[position()=last()-3]'
        )

        # Obtiene el EPS del año fiscal anterior y lo almacena en eps_last_fiscal_year.
        while not eps_last_fiscal_year:
            eps_last_fiscal_year = self.pyppeteer_extensions.get_text(
                '(//div[text()="Diluted EPS"]/parent::div/parent::div/parent::div/parent::div/div)[position()=last()-3]'
            )

        # Espera 2 segundos para permitir que la interfaz cargue.
        time.sleep(2)

        # Hace clic en la pestaña "Solvency" (Solvencia) para obtener información adicional.
        self.pyppeteer_extensions.click('//span[text()="Solvency"]')

        # Espera a que aparezca el elemento que contiene la relación de pasivos totales sobre activos totales.
        self.pyppeteer_extensions.wait_for_element(
            '(//div[text()="Total Liabilities / Total Assets"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
        )

        # Obtiene el valor de la relación de pasivos totales sobre activos totales y lo almacena en total_liabilities_total_assets.
        while not total_liabilities_total_assets:
            total_liabilities_total_assets = self.pyppeteer_extensions.get_text(
                '(//div[text()="Total Liabilities / Total Assets"]/parent::div/parent::div/parent::div/parent::div/div)[last()]'
            )

        # Espera 3 segundos para permitir que la interfaz cargue antes de continuar.
        time.sleep(3)

        # Elimina las comas de los valores de las métricas financieras para permitir su conversión a tipo float.
        pe_ratio = pe_ratio.replace(",", "")
        pb_ratio = pb_ratio.replace(",", "")
        stock_price = stock_price.replace(",", "")
        current_ratio = current_ratio.replace(",", "")
        return_on_equity = return_on_equity.replace(",", "")
        asset_turnover = asset_turnover.replace(",", "")
        eps_current_fiscal_year = eps_current_fiscal_year.replace(",", "")
        eps_last_fiscal_year = eps_last_fiscal_year.replace(",", "")
        total_liabilities_total_assets = total_liabilities_total_assets.replace(",", "")

        # Elimina la "x" de las métricas de rotación de activos y relación corriente.
        asset_turnover = asset_turnover.replace("x", "")
        current_ratio = current_ratio.replace("x", "")

        # Elimina el símbolo de porcentaje de la rentabilidad sobre el capital y la relación de pasivos sobre activos.
        return_on_equity = return_on_equity.replace("%", "")
        total_liabilities_total_assets = total_liabilities_total_assets.replace("%", "")

        # Convierte las métricas a tipo float, verificando si su valor es "Upgrade" antes de la conversión.
        # Si el valor es "Upgrade", lo deja como está.
        pe_ratio = float(pe_ratio) if pe_ratio != "Upgrade" else pe_ratio
        pb_ratio = float(pb_ratio) if pb_ratio != "Upgrade" else pb_ratio
        stock_price = float(stock_price) if stock_price != "Upgrade" else stock_price
        current_ratio = (
            float(current_ratio) if current_ratio != "Upgrade" else current_ratio
        )
        return_on_equity = (
            float(return_on_equity)
            if return_on_equity != "Upgrade"
            else return_on_equity
        )
        asset_turnover = (
            float(asset_turnover) if asset_turnover != "Upgrade" else asset_turnover
        )
        eps_current_fiscal_year = (
            float(eps_current_fiscal_year)
            if eps_current_fiscal_year != "Upgrade"
            else eps_current_fiscal_year
        )
        eps_last_fiscal_year = (
            float(eps_last_fiscal_year)
            if eps_last_fiscal_year != "Upgrade"
            else eps_last_fiscal_year
        )
        total_liabilities_total_assets = (
            float(total_liabilities_total_assets)
            if total_liabilities_total_assets != "Upgrade"
            else total_liabilities_total_assets
        )

        # Llamar al método para escribir los datos recopilados en un archivo Excel.
        self.write_to_excel(
            char,
            pe_ratio,
            pb_ratio,
            stock_price,
            current_ratio,
            return_on_equity,
            asset_turnover,
            eps_current_fiscal_year,
            eps_last_fiscal_year,
            total_liabilities_total_assets,
        )

    # ----------------------------------------------------------------------------------------------

    """
    self,
    char,  # Caracter ASCII que representa la columna en la hoja de Excel donde se escribirán los datos.
    pe_ratio,  # Relación precio/ganancias de la acción.
    pb_ratio,  # Relación precio/valor contable de la acción.
    stock_price,  # Precio actual de la acción.
    current_ratio,  # Relación corriente, que mide la capacidad de una empresa para pagar sus deudas a corto plazo.
    return_on_equity,  # Retorno sobre el capital (ROE), que mide la rentabilidad sobre los fondos propios.
    asset_turnover,  # Rotación de activos, que mide la eficiencia en la utilización de los activos para generar ventas.
    eps_current_fiscal_year,  # Ganancias por acción (EPS) del año fiscal actual.
    eps_last_fiscal_year,  # Ganancias por acción (EPS) del año fiscal anterior.
    total_liabilities_total_assets,  # Proporción de pasivos totales sobre activos totales, que mide la solvencia de la empresa.
    """

    def write_to_excel(
        self,
        char,
        pe_ratio,
        pb_ratio,
        stock_price,
        current_ratio,
        return_on_equity,
        asset_turnover,
        eps_current_fiscal_year,
        eps_last_fiscal_year,
        total_liabilities_total_assets,
    ):
        try:
            # Intenta cargar el libro de trabajo de Excel especificado en `self.file_name`.
            workbook = load_workbook(filename=self.file_name)

            # Obtiene la hoja activa del libro de trabajo.
            sheet = workbook.active

            # Convierte el valor ASCII de `char` a su correspondiente letra de columna.
            column = chr(char)

            # Escribe los datos en las celdas especificadas de la columna seleccionada.
            # Relación P/E en la celda correspondiente.
            sheet[column + "3"] = pe_ratio

            # EPS del año anterior.
            sheet[column + "4"] = eps_last_fiscal_year

            # EPS del año actual.
            sheet[column + "5"] = eps_current_fiscal_year

            # Relación P/B.
            sheet[column + "8"] = pb_ratio

            # ROE.
            sheet[column + "9"] = return_on_equity

            # Relación corriente.
            sheet[column + "10"] = current_ratio

            # Proporción de pasivos totales sobre activos totales.
            sheet[column + "11"] = total_liabilities_total_assets

            # Rotación de activos.
            sheet[column + "12"] = asset_turnover

            # Precio de la acción.
            sheet[column + "15"] = stock_price

            # Guarda los cambios en el archivo de Excel.
            workbook.save(self.file_name)
        except Exception as e:
            # Captura y muestra un mensaje de error si ocurre un problema al escribir en el archivo de Excel.
            print(f"Error writing to Excel file: {e}")

    # ----------------------------------------------------------------------------------------------

    def get_ticker_list(self):
        try:
            # Intenta cargar el libro de trabajo de Excel especificado en `self.file_name`.
            workbook = load_workbook(filename=self.file_name)

            # Obtiene la hoja activa del libro de trabajo.
            sheet = workbook.active

            # Crea una lista de tickers extrayendo los valores de las celdas C1, D1, E1 y F1.
            tickers_list = [
                str(sheet["C1"].value),
                str(sheet["D1"].value),
                str(sheet["E1"].value),
                str(sheet["F1"].value),
            ]

            # Devuelve la lista de tickers.
            return tickers_list
        except FileNotFoundError:
            # Si el archivo de Excel no se encuentra, imprime un mensaje de error.
            print(f"Excel file not found: {self.file_name}")
            return []
        except Exception as e:
            # Captura cualquier otra excepción y muestra un mensaje de error.
            print(f"Error reading Excel file: {e}")
            return []


# ----------------------------------------------------------------------------------------------

# Verifica si el script se está ejecutando directamente (no importado como un módulo).
if __name__ == "__main__":
    # Carga las variables de entorno desde el archivo .env al entorno de ejecución actual.
    load_dotenv()

    # Define la ruta al archivo JSON de configuración que será usado por el bot.
    json_file_path = "C:\\stocks_bot\\config.json"

    # Registra el tiempo de inicio para medir el tiempo total de ejecución del bot.
    start_time = time.perf_counter()

    # Instancia y ejecuta el bot `StocksBot`, pasando el archivo de configuración como argumento.
    bot = StocksBot(json_file_path)

    # Registra el tiempo de finalización de la ejecución del bot.
    end_time = time.perf_counter()

    # Calcula el tiempo total que ha tardado el bot en ejecutarse.
    total_time = end_time - start_time

    # Imprime el tiempo total de ejecución del bot, formateado con dos decimales.
    print(f"El bot ha tardado {total_time:.2f} segundos en ejecutarse.")
